from __future__ import annotations

import json
import sys
import unittest
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from invoice_automation.pipeline import run_pipeline
from invoice_automation.reporting import (
    build_excel_report,
    build_json_report,
)


class PipelineAndReportTests(unittest.TestCase):
    def test_pipeline_calculates_values_and_overdue_status(self) -> None:
        data = pd.DataFrame(
            [
                {
                    "invoice_id": "FV-PLN",
                    "client_name": "Client PLN",
                    "issue_date": "2026-07-01",
                    "due_date": "2026-07-10",
                    "currency": "PLN",
                    "net_amount": 100,
                    "vat_rate": 23,
                    "payment_status": "paid",
                },
                {
                    "invoice_id": "FV-EUR",
                    "client_name": "Client EUR",
                    "issue_date": "2026-06-01",
                    "due_date": "2026-07-01",
                    "currency": "EUR",
                    "net_amount": 100,
                    "vat_rate": 23,
                    "payment_status": "unpaid",
                },
            ]
        )

        result = run_pipeline(
            data,
            reporting_date="2026-07-30",
            use_api=False,
        )

        self.assertEqual(result.summary["valid_invoices"], 2)
        self.assertEqual(result.summary["invalid_rows"], 0)
        self.assertEqual(result.summary["overdue_invoices"], 1)
        self.assertEqual(result.summary["total_gross_pln"], 651.9)
        eur = result.valid_invoices.loc[
            result.valid_invoices["currency"].eq("EUR")
        ].iloc[0]
        self.assertEqual(eur["gross_amount_pln"], 528.9)
        self.assertEqual(eur["days_overdue"], 29)
        self.assertEqual(eur["rate_source"], "demo_fallback")

    def test_generated_excel_contains_expected_sheets(self) -> None:
        sample = pd.read_csv(PROJECT_ROOT / "data" / "sample_invoices.csv")
        result = run_pipeline(
            sample,
            reporting_date="2026-07-30",
            use_api=False,
        )

        workbook = load_workbook(BytesIO(build_excel_report(result)))

        self.assertEqual(
            workbook.sheetnames,
            [
                "Summary",
                "Valid invoices",
                "All rows",
                "Validation issues",
                "By currency",
                "By client",
                "Exchange rates",
            ],
        )
        self.assertEqual(
            workbook["Summary"]["A1"].value,
            "Invoice Automation Report",
        )
        self.assertEqual(
            workbook["Summary"]["A4"].value,
            "Metric",
        )

    def test_json_report_is_valid_and_contains_summary(self) -> None:
        sample = pd.read_csv(PROJECT_ROOT / "data" / "sample_invoices.csv")
        result = run_pipeline(
            sample,
            reporting_date="2026-07-30",
            use_api=False,
        )

        payload = json.loads(build_json_report(result).decode("utf-8"))

        self.assertEqual(payload["reporting_date"], "2026-07-30")
        self.assertEqual(payload["summary"]["input_rows"], len(sample))
        self.assertIn("validation_issues", payload)
        self.assertIn("exchange_rates", payload)


if __name__ == "__main__":
    unittest.main()
