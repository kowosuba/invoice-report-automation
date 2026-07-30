from __future__ import annotations

import json
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .pipeline import PipelineResult


HEADER_FILL = PatternFill("solid", fgColor="27866C")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBTLE_FILL = PatternFill("solid", fgColor="EAF5F1")
ERROR_FILL = PatternFill("solid", fgColor="FCE8E6")
OVERDUE_FILL = PatternFill("solid", fgColor="FFF1CC")


def _records(frame: pd.DataFrame) -> list[dict]:
    clean = frame.copy()
    for column in clean.columns:
        if pd.api.types.is_datetime64_any_dtype(clean[column]):
            clean[column] = clean[column].dt.strftime("%Y-%m-%d")
    clean = clean.astype(object).where(pd.notna(clean), None)
    return clean.to_dict(orient="records")


def build_json_report(result: PipelineResult) -> bytes:
    payload = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "reporting_date": result.reporting_date,
        "summary": result.summary,
        "exchange_rates": {
            code: {
                "rate_to_pln": rate.rate_to_pln,
                "source": rate.source,
                "effective_date": rate.effective_date,
            }
            for code, rate in result.exchange_rates.items()
        },
        "warnings": result.warnings,
        "validation_issues": _records(result.issues),
        "valid_invoices": _records(result.valid_invoices),
    }
    return json.dumps(
        payload,
        ensure_ascii=False,
        indent=2,
        default=str,
    ).encode("utf-8")


def build_csv_report(result: PipelineResult) -> bytes:
    return result.processed_data.to_csv(index=False).encode("utf-8-sig")


def _style_excel(
    workbook_bytes: BytesIO,
    result: PipelineResult,
) -> bytes:
    workbook_bytes.seek(0)
    workbook = load_workbook(workbook_bytes)

    for worksheet in workbook.worksheets:
        worksheet.sheet_view.showGridLines = False
        header_row = 4 if worksheet.title == "Summary" else 1
        worksheet.freeze_panes = f"A{header_row + 1}"
        worksheet.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(worksheet.max_column)}"
            f"{worksheet.max_row}"
        )
        worksheet.row_dimensions[header_row].height = 24

        for cell in worksheet[header_row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center")

        for column_index in range(1, worksheet.max_column + 1):
            values = [
                str(worksheet.cell(row, column_index).value or "")
                for row in range(1, worksheet.max_row + 1)
            ]
            width = min(max(max(map(len, values)) + 2, 11), 42)
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

        header_map = {
            str(worksheet.cell(header_row, column).value): column
            for column in range(1, worksheet.max_column + 1)
        }
        for row in worksheet.iter_rows(min_row=header_row + 1):
            for cell in row:
                header = str(worksheet.cell(header_row, cell.column).value or "")
                if "amount" in header or "rate_to_pln" in header:
                    cell.number_format = "#,##0.00"
                if "date" in header and cell.value:
                    cell.number_format = "yyyy-mm-dd"
                cell.alignment = Alignment(vertical="top")

        if worksheet.max_row > header_row:
            table_name = (
                worksheet.title.replace(" ", "").replace("-", "") + "Table"
            )
            table = Table(
                displayName=table_name,
                ref=(
                    f"A{header_row}:"
                    f"{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
                ),
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)

        if worksheet.max_row > header_row and "is_valid" in header_map:
            column = get_column_letter(header_map["is_valid"])
            worksheet.conditional_formatting.add(
                f"{column}{header_row + 1}:{column}{worksheet.max_row}",
                CellIsRule(
                    operator="equal",
                    formula=["FALSE"],
                    fill=ERROR_FILL,
                ),
            )
        if worksheet.max_row > header_row and "is_overdue" in header_map:
            column = get_column_letter(header_map["is_overdue"])
            worksheet.conditional_formatting.add(
                f"{column}{header_row + 1}:{column}{worksheet.max_row}",
                CellIsRule(
                    operator="equal",
                    formula=["TRUE"],
                    fill=OVERDUE_FILL,
                ),
            )

    summary_sheet = workbook["Summary"]
    summary_sheet.merge_cells("A1:D1")
    summary_sheet["A1"] = "Invoice Automation Report"
    summary_sheet["A1"].fill = HEADER_FILL
    summary_sheet["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    summary_sheet["A1"].alignment = Alignment(vertical="center")
    summary_sheet.row_dimensions[1].height = 34

    summary_sheet["A2"] = "Reporting date"
    summary_sheet["B2"] = result.reporting_date
    summary_sheet["C2"] = "Valid / invalid rows"
    summary_sheet["D2"] = (
        f"{result.summary['valid_invoices']} / "
        f"{result.summary['invalid_rows']}"
    )
    for row in summary_sheet["A2:D2"]:
        for cell in row:
            cell.fill = SUBTLE_FILL
            cell.font = Font(bold=True)

    summary_sheet.merge_cells("A3:D3")
    summary_sheet["A3"] = (
        "CSV validation, currency conversion, overdue detection and reporting"
    )
    summary_sheet["A3"].font = Font(color="4E5955", italic=True)
    summary_sheet["A3"].alignment = Alignment(vertical="center")
    summary_sheet.column_dimensions["A"].width = 27
    summary_sheet.column_dimensions["B"].width = 18
    summary_sheet.column_dimensions["C"].width = 23
    summary_sheet.column_dimensions["D"].width = 16

    for row in range(5, summary_sheet.max_row + 1):
        label = str(summary_sheet.cell(row, 1).value or "")
        if "PLN" in label:
            summary_sheet.cell(row, 2).number_format = '#,##0.00 "PLN"'

    styled = BytesIO()
    workbook.save(styled)
    return styled.getvalue()


def build_excel_report(result: PipelineResult) -> bytes:
    stream = BytesIO()
    metric_labels = {
        "input_rows": "Input rows",
        "valid_invoices": "Valid invoices",
        "invalid_rows": "Invalid rows",
        "validation_issues": "Validation issues",
        "unique_clients": "Unique clients",
        "overdue_invoices": "Overdue invoices",
        "total_gross_pln": "Total gross (PLN)",
        "overdue_gross_pln": "Overdue gross (PLN)",
    }
    summary_frame = pd.DataFrame(
        [
            {
                "Metric": metric_labels.get(key, key),
                "Value": value,
            }
            for key, value in result.summary.items()
        ]
    )
    rates_frame = pd.DataFrame(
        [
            {
                "currency": code,
                "rate_to_pln": rate.rate_to_pln,
                "source": rate.source,
                "effective_date": rate.effective_date,
                "source_url": (
                    "https://api.nbp.pl/"
                    if rate.source == "NBP table A"
                    else ""
                ),
            }
            for code, rate in sorted(result.exchange_rates.items())
        ]
    )

    with pd.ExcelWriter(stream, engine="openpyxl") as writer:
        summary_frame.to_excel(
            writer,
            sheet_name="Summary",
            index=False,
            startrow=3,
        )
        result.valid_invoices.to_excel(
            writer,
            sheet_name="Valid invoices",
            index=False,
        )
        result.processed_data.to_excel(
            writer,
            sheet_name="All rows",
            index=False,
        )
        result.issues.to_excel(
            writer,
            sheet_name="Validation issues",
            index=False,
        )
        result.by_currency.to_excel(
            writer,
            sheet_name="By currency",
            index=False,
        )
        result.by_client.to_excel(
            writer,
            sheet_name="By client",
            index=False,
        )
        rates_frame.to_excel(
            writer,
            sheet_name="Exchange rates",
            index=False,
        )

    return _style_excel(stream, result)


def write_reports(
    result: PipelineResult,
    output_directory: str | Path,
) -> dict[str, Path]:
    output = Path(output_directory)
    output.mkdir(parents=True, exist_ok=True)

    paths = {
        "excel": output / "invoice_report.xlsx",
        "csv": output / "processed_invoices.csv",
        "json": output / "summary_report.json",
        "issues": output / "validation_issues.csv",
        "log": output / "run.log",
    }
    paths["excel"].write_bytes(build_excel_report(result))
    paths["csv"].write_bytes(build_csv_report(result))
    paths["json"].write_bytes(build_json_report(result))
    result.issues.to_csv(paths["issues"], index=False, encoding="utf-8-sig")

    log_lines = [
        f"Reporting date: {result.reporting_date}",
        *(f"{key}: {value}" for key, value in result.summary.items()),
        *(f"WARNING: {warning}" for warning in result.warnings),
    ]
    paths["log"].write_text("\n".join(log_lines) + "\n", encoding="utf-8")
    return paths
